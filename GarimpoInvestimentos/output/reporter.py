import csv
import logging
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from GarimpoInvestimentos.core.paths import OUTPUT_DIR

_log = logging.getLogger("previsao_cripto.reporter")


def export_results(resultados: list[dict]):
    # UTC como o resto do projeto (C7; previsões são carimbadas em UTC desde
    # 2026-07-07 — nome de arquivo em hora local criava skew de até 3h no par)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    csv_filename = str(OUTPUT_DIR / f"garimpo_resultados_{timestamp}.csv")
    xlsx_filename = str(OUTPUT_DIR / f"garimpo_resultados_{timestamp}.xlsx")

    # CSV
    fieldnames = ["Ativo", "Sentimento", "Score", "Resumo", "Data", "Preço USD"]
    with open(csv_filename, mode="w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for r in resultados:
            writer.writerow(
                {
                    "Ativo": r.get("ativo", "").upper(),
                    "Sentimento": r.get("sentimento", ""),
                    "Score": round(r.get("score", 0), 2),
                    "Resumo": r.get("resumo", ""),
                    "Data": r.get("data", ""),
                    "Preço USD": round(r.get("price_usd", 0), 2),
                }
            )

    # XLSX
    wb = Workbook()
    ws = wb.active
    assert isinstance(ws, Worksheet)  # Workbook() sem read_only sempre dá Worksheet
    ws.title = "Análises"

    headers = ["Ativo", "Sentimento", "Score", "Resumo", "Data", "Preço USD"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="B00020")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in resultados:
        ws.append(
            [
                r.get("ativo", "").upper(),
                r.get("sentimento", ""),
                round(r.get("score", 0), 2),
                r.get("resumo", ""),
                r.get("data", ""),
                round(r.get("price_usd", 0), 2),
            ]
        )

    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    score_col = "C"
    start_row = 2
    end_row = len(resultados) + 1

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFF3B0")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    ws.conditional_formatting.add(
        f"{score_col}{start_row}:{score_col}{end_row}",
        CellIsRule(operator="greaterThan", formula=["70"], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"{score_col}{start_row}:{score_col}{end_row}",
        CellIsRule(operator="between", formula=["50", "70"], fill=yellow_fill),
    )
    ws.conditional_formatting.add(
        f"{score_col}{start_row}:{score_col}{end_row}",
        CellIsRule(operator="lessThan", formula=["50"], fill=red_fill),
    )

    chart = BarChart()
    chart.title = "Pontuação de Oportunidade (Score)"
    chart.y_axis.title = "Ativos"
    chart.x_axis.title = "Score"
    chart.style = 13
    chart.type = "bar"
    chart.width = 15
    chart.height = 6

    data = Reference(ws, min_col=3, min_row=1, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    chart_anchor_row = end_row + 3
    ws.add_chart(chart, f"A{chart_anchor_row}")  # pyright: ignore[reportCallIssue] — stub do openpyxl confunde Worksheet/Chartsheet.add_chart

    wb.save(xlsx_filename)

    _log.info("Resultados exportados: CSV -> %s", csv_filename)
    _log.info(
        "Resultados exportados: XLSX -> %s (com grafico e formatacao condicional)", xlsx_filename
    )
